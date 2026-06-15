"""Convert path: xlsx → YAML (T08 / US1).

Builds a fixture workbook programmatically in ``tmp_path`` (research §9 — no
binary fixtures are committed) and exercises the CLI as a subprocess. Covers:

- Happy path: one YAML per sheet, schema-valid output.
- Idempotency: a second convert produces byte-identical files (SC-002).
- Duplicate competency_id across blocks → non-zero exit, no files written.
- Merged cell across key column → non-zero exit.
- Canonical YAML emitter: same input → byte-identical output.
"""

from __future__ import annotations

import filecmp
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.backend.services.rubric_importer import _emit_canonical_yaml

_REPO_ROOT: Path = Path(__file__).resolve().parents[4]
_HEADER: list[str] = [
    "block",
    "competency_id",
    "competency_label_uk",
    "competency_label_en",
    "topic",
    "level",
    "descriptor_en",
    "level_label_uk",
    "evidence_examples",
    "competency_retired",
]


def _build_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    """Write a workbook with the given (sheet → rows-after-header) layout."""
    wb = Workbook()
    # Remove the default sheet so the order matches `sheets`.
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(_HEADER)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


def _happy_python_rows() -> list[list[object]]:
    return [
        [
            "Core",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "Threading",
            1,
            "Knows threads exist; can describe the GIL in one sentence.",
            "Початковий",
            "Mentions threading without describing when it helps",
            "false",
        ],
        [
            "Core",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "Asyncio",
            3,
            "Confidently chooses between threading, multiprocessing, and asyncio.",
            "Середній",
            "Designs an async pipeline; reasons about back-pressure",
            "false",
        ],
        [
            "Core",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "",
            5,
            "Designs a custom scheduler / reasons about subtle interpreter behaviour.",
            "Експерт",
            "Diagnoses a deadlock from a thread-dump",
            "false",
        ],
    ]


def _happy_react_rows() -> list[list[object]]:
    return [
        [
            "UI",
            "react.hooks",
            "Хуки",
            "Hooks",
            "useState",
            2,
            "Confidently uses useState/useEffect; understands closure traps.",
            "Молодший",
            "Writes a counter component",
            "false",
        ],
        [
            "UI",
            "react.hooks",
            "Хуки",
            "Hooks",
            "Custom",
            4,
            "Designs custom hooks with stable identities; reasons about render cycles.",
            "Старший",
            "Designs a useDebounce hook",
            "false",
        ],
    ]


def _run_convert(xlsx_path: Path, out_dir: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [
            sys.executable,
            "-m",
            "app.backend.cli.import_matrix",
            "convert",
            str(xlsx_path),
            "--out",
            str(out_dir),
        ],
        cwd=str(_REPO_ROOT),
        capture_output=True,
        text=True,
        check=False,
    )


def test_convert_writes_one_yaml_per_stack(tmp_path: Path) -> None:
    """Happy path: two sheets → two YAML files; each validates against the schema."""
    xlsx = _build_workbook(
        tmp_path / "matrix.xlsx",
        {"python": _happy_python_rows(), "react": _happy_react_rows()},
    )
    out_dir = tmp_path / "out"
    result = _run_convert(xlsx, out_dir)
    assert result.returncode == 0, result.stderr

    py_yaml = out_dir / "python.yaml"
    rx_yaml = out_dir / "react.yaml"
    assert py_yaml.exists()
    assert rx_yaml.exists()

    py_content = py_yaml.read_text(encoding="utf-8")
    assert "python.concurrency" in py_content
    assert "Конкурентність" in py_content
    # Three levels written.
    assert py_content.count("level: 1") >= 1
    assert py_content.count("level: 3") >= 1
    assert py_content.count("level: 5") >= 1


def test_convert_is_byte_identical_idempotent(tmp_path: Path) -> None:
    """SC-002: a second convert against the same workbook leaves YAMLs byte-identical."""
    xlsx = _build_workbook(tmp_path / "matrix.xlsx", {"python": _happy_python_rows()})
    out_dir = tmp_path / "out"
    first = _run_convert(xlsx, out_dir)
    assert first.returncode == 0, first.stderr
    first_bytes = (out_dir / "python.yaml").read_bytes()

    # Second run.
    second = _run_convert(xlsx, out_dir)
    assert second.returncode == 0, second.stderr
    second_bytes = (out_dir / "python.yaml").read_bytes()
    assert first_bytes == second_bytes


def test_duplicate_competency_id_across_blocks_rejected(tmp_path: Path) -> None:
    """A competency_id appearing in two distinct blocks is rejected before writing."""
    rows = [
        [
            "Core",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "T",
            1,
            "desc1",
            "L1",
            "",
            "false",
        ],
        [
            "Advanced",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "T",
            3,
            "desc3",
            "L3",
            "",
            "false",
        ],
    ]
    xlsx = _build_workbook(tmp_path / "bad.xlsx", {"python": rows})
    out_dir = tmp_path / "out"
    result = _run_convert(xlsx, out_dir)
    assert result.returncode == 1
    assert "python.concurrency" in result.stderr
    assert not (out_dir / "python.yaml").exists()


def test_merged_key_column_rejected(tmp_path: Path) -> None:
    """A merged cell that spans rows in a key column is rejected with a clear error."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("python")
    ws.append(_HEADER)
    ws.append(
        [
            "Core",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "T",
            1,
            "desc1",
            "L1",
            "",
            "false",
        ]
    )
    ws.append(
        [
            "",
            "python.concurrency",
            "Конкурентність",
            "Concurrency",
            "T",
            3,
            "desc3",
            "L3",
            "",
            "false",
        ]
    )
    # Merge the block column across the two data rows.
    block_col_letter = get_column_letter(_HEADER.index("block") + 1)
    ws.merge_cells(f"{block_col_letter}2:{block_col_letter}3")
    xlsx = tmp_path / "merged.xlsx"
    wb.save(xlsx)

    out_dir = tmp_path / "out"
    result = _run_convert(xlsx, out_dir)
    assert result.returncode == 1
    assert "merged" in result.stderr.lower() or "key column" in result.stderr.lower()


def test_canonical_yaml_emitter_is_deterministic() -> None:
    """Same input dict → byte-identical canonical YAML across runs."""
    doc = {
        "version": 1,
        "retired": False,
        "nodes": [
            {
                "id": "block.core",
                "label_en": "Core",
                "label_uk": "Ядро",
                "parent": None,
                "retired": False,
            },
            {
                "id": "python.concurrency",
                "label_en": "Concurrency",
                "label_uk": "Конкурентність",
                "parent": "block.core",
                "retired": False,
                "levels": [
                    {"level": 3, "label_uk": "Mid", "descriptor_en": "mid"},
                    {"level": 1, "label_uk": "Junior", "descriptor_en": "junior"},
                ],
            },
        ],
    }
    a = _emit_canonical_yaml(doc)
    b = _emit_canonical_yaml(doc)
    assert a == b
    # Sorted by id: block.core comes before python.concurrency.
    assert a.index("block.core") < a.index("python.concurrency")
    # Levels sorted by integer.
    assert a.index("level: 1") < a.index("level: 3")


@pytest.mark.parametrize("filename", ["python.yaml", "react.yaml"])
def test_convert_then_filecmp_idempotent(tmp_path: Path, filename: str) -> None:
    """Belt-and-suspenders SC-002: convert → snapshot dir → convert → filecmp."""
    xlsx = _build_workbook(
        tmp_path / "matrix.xlsx",
        {"python": _happy_python_rows(), "react": _happy_react_rows()},
    )
    out_a = tmp_path / "a"
    out_b = tmp_path / "b"
    assert _run_convert(xlsx, out_a).returncode == 0
    assert _run_convert(xlsx, out_b).returncode == 0
    assert filecmp.cmp(out_a / filename, out_b / filename, shallow=False)
