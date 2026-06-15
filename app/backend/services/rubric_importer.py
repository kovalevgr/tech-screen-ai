"""Rubric matrix importer (T08).

Two orthogonal capabilities sharing a single canonical-YAML emitter (research §7):

- :meth:`RubricImporter.convert` — read an Excel workbook (one sheet per stack),
  validate the matrix-format contract, and emit one canonical YAML per stack
  into ``configs/rubric/<stack-id>.yaml``. Idempotent: a second run on the same
  workbook leaves the YAMLs byte-identical (FR-004 / SC-002).
- :meth:`RubricImporter.seed` — read every ``configs/rubric/*.yaml``, validate
  it against ``docs/contracts/rubric.schema.json``, compute a SHA-256 payload
  hash, and reconcile the database with §4 immutability: new
  ``rubric_tree_version`` row per content change, zero edits on prior-version
  rows, one ``audit_log`` row per new version.

§4 / ADR-018 immutability is structural: prior-version rows are NEVER updated
or deleted. The seed path inserts new rows under a new version id when the
payload hash differs; identical hash → no-op (research §3).

§3 carve-out: the rubric-tree tables (``stack`` / ``competency_block`` /
``competency`` / ``topic`` / ``level``) are NOT in the append-only set —
mutation across versions is the design. ``audit_log`` IS in the §3 set; this
importer only INSERTs there (one row per new version, FR-010).
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Any, cast
from uuid import UUID

import asyncpg
import yaml
from jsonschema import Draft202012Validator

_log = logging.getLogger(__name__)

# Repo-relative defaults; tests override.
_REPO_ROOT: Path = Path(__file__).resolve().parents[3]
_SCHEMA_PATH: Path = _REPO_ROOT / "docs" / "contracts" / "rubric.schema.json"
_DEFAULT_RUBRIC_DIR: Path = _REPO_ROOT / "configs" / "rubric"

# Stable id regex — same one the JSON Schema enforces.
_STABLE_ID_RE: re.Pattern[str] = re.compile(r"^[a-z][a-z0-9_]*(\.[a-z][a-z0-9_]*)*$")

# Workbook header — the canonical column names from docs/contracts/matrix-format.md.
_REQUIRED_HEADERS: tuple[str, ...] = (
    "block",
    "competency_id",
    "competency_label_uk",
    "competency_label_en",
    "level",
    "descriptor_en",
    "level_label_uk",
)
_OPTIONAL_HEADERS: tuple[str, ...] = (
    "topic",
    "evidence_examples",
    "competency_retired",
)
_KEY_COLUMNS: tuple[str, ...] = ("block", "competency_id", "topic", "level")

# Advisory-lock id for the seed transaction (research §6).
_SEED_ADVISORY_LOCK_ID: int = 987654321
_NOTIFY_PG_CHANNEL: str = "rubric_changed"


# ---------------------------------------------------------------------------
# Typed errors
# ---------------------------------------------------------------------------


class RubricImporterError(Exception):
    """Base for all importer errors. CLI prints these as one-line stderr messages."""


class UnknownColumnError(RubricImporterError):
    """A required header column is missing from a sheet."""


class EmptyRequiredCellError(RubricImporterError):
    """A required cell is empty."""


class DuplicateCompetencyIdError(RubricImporterError):
    """The same competency_id appears in two rows of the same sheet."""


class MergedKeyColumnError(RubricImporterError):
    """A merged cell spans a key column (block/competency_id/topic/level)."""


class SchemaViolationError(RubricImporterError):
    """A YAML file violates docs/contracts/rubric.schema.json."""


class RenameForbiddenError(RubricImporterError):
    """A stable id active in the prior version disappeared without being retired (FR-009)."""


# ---------------------------------------------------------------------------
# Public dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class SeedResult:
    """Outcome of a single :meth:`RubricImporter.seed` invocation."""

    noop: bool
    """True when the payload hash matched the latest version (no rows written)."""

    new_version_id: UUID | None
    """The new ``rubric_tree_version.id`` when a new version was created; None on no-op."""

    new_payload_hash: str
    """The hash of the current payload (always populated)."""

    rows_inserted: int
    """Total rows inserted across rubric tree tables + the one audit_log row."""


# ---------------------------------------------------------------------------
# Canonical YAML + hash helpers (research §2/§3/§7)
# ---------------------------------------------------------------------------


def _emit_canonical_yaml(doc: dict[str, Any]) -> str:
    """Serialise ``doc`` to canonical YAML (byte-identical across reruns).

    Pre-sorts ``nodes`` by ``id`` and each node's ``levels`` by ``level``
    (research §2 — ``yaml.safe_dump(sort_keys=True)`` does not reorder list
    contents).
    """
    nodes = sorted(doc.get("nodes") or [], key=lambda n: cast("str", n["id"]))
    sorted_nodes: list[dict[str, Any]] = []
    for n in nodes:
        node = dict(n)
        if "levels" in node and node["levels"] is not None:
            node["levels"] = sorted(node["levels"], key=lambda lvl: cast("int", lvl["level"]))
        sorted_nodes.append(node)
    canonical = dict(doc)
    canonical["nodes"] = sorted_nodes
    out = StringIO()
    yaml.safe_dump(
        canonical,
        stream=out,
        sort_keys=True,
        default_flow_style=False,
        allow_unicode=True,
        indent=2,
        width=120,
        line_break="\n",
    )
    return out.getvalue()


def _compute_payload_hash(yaml_dir: Path) -> str:
    """SHA-256 hex of the canonical bytes of every yaml file, filename-sorted (research §3)."""
    sha = hashlib.sha256()
    files = sorted(yaml_dir.glob("*.yaml"))
    for i, path in enumerate(files):
        if i > 0:
            sha.update(b"\x00")
        sha.update(path.read_bytes())
    return sha.hexdigest()


def _load_schema(schema_path: Path = _SCHEMA_PATH) -> dict[str, Any]:
    return cast("dict[str, Any]", json.loads(schema_path.read_text(encoding="utf-8")))


def _validate_yaml(doc: dict[str, Any], schema: dict[str, Any], origin: str) -> None:
    validator = Draft202012Validator(schema)
    errors = list(validator.iter_errors(doc))
    if errors:
        first = errors[0]
        path = ".".join(str(p) for p in first.absolute_path) or "<root>"
        raise SchemaViolationError(f"{origin}: ${path}: {first.message}")


def _asyncpg_dsn(sqlalchemy_dsn: str) -> str:
    return sqlalchemy_dsn.replace("+asyncpg", "")


def _norm(value: str | None) -> str:
    """Strip + NFC-normalise a cell value (research §8)."""
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value).strip())


def _slugify(name: str) -> str:
    """Slugify a free-form block name to a stable-id segment.

    Produces a snake_case ASCII slug acceptable to the stable-id regex.
    """
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "block"


# ---------------------------------------------------------------------------
# Convert path (xlsx → YAML)
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class _ParsedLevel:
    level: int
    label_uk: str
    descriptor_en: str
    evidence_examples_en: list[str] = field(default_factory=list)


@dataclass(slots=True)
class _ParsedCompetency:
    id: str
    block_slug: str
    block_label: str
    label_uk: str
    label_en: str
    topic: str
    retired: bool = False
    levels: dict[int, _ParsedLevel] = field(default_factory=dict)


def _read_sheet(sheet: Any, sheet_name: str) -> dict[str, _ParsedCompetency]:
    """Return mapping competency_id → parsed competency for one sheet."""
    # Header row.
    rows = sheet.iter_rows(values_only=False)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise UnknownColumnError(f"sheet '{sheet_name}': empty workbook") from exc
    headers: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = _norm(cell.value)
        if name:
            headers[name] = idx
    for col in _REQUIRED_HEADERS:
        if col not in headers:
            raise UnknownColumnError(
                f"sheet '{sheet_name}': required column '{col}' missing from header"
            )

    # Merged-cell check on key columns.
    merged_ranges = list(sheet.merged_cells.ranges)
    for col in _KEY_COLUMNS:
        if col not in headers:
            continue
        col_idx = headers[col]
        for mr in merged_ranges:
            # mr.min_col / mr.max_col are 1-indexed; openpyxl cell.column is also 1-indexed.
            # cell index in the row tuple is 0-indexed → col_idx + 1 == column number.
            col_number = col_idx + 1
            if mr.min_col <= col_number <= mr.max_col and mr.min_row != mr.max_row:
                raise MergedKeyColumnError(
                    f"sheet '{sheet_name}': merged cells {mr.coord} span key column '{col}'"
                )

    def cell_at(row: tuple[Any, ...], name: str) -> str:
        idx = headers.get(name)
        if idx is None or idx >= len(row):
            return ""
        return _norm(row[idx].value)

    parsed: dict[str, _ParsedCompetency] = {}
    for row_idx, row in enumerate(rows, start=2):

        def cell(name: str, _row: tuple[Any, ...] = row) -> str:
            return cell_at(_row, name)

        block = cell("block")
        competency_id = cell("competency_id")
        # Tolerate fully-empty trailing rows (research §8).
        if not block and not competency_id:
            # Check if entire row is empty
            if not any(_norm(c.value) for c in row):
                continue
        if not competency_id:
            # A row with empty competency_id but other content — treat as section break (skip).
            continue
        if not _STABLE_ID_RE.match(competency_id):
            raise EmptyRequiredCellError(
                f"sheet '{sheet_name}' row {row_idx}: competency_id "
                f"'{competency_id}' is not a valid stable id"
            )
        for required in (
            "block",
            "competency_label_uk",
            "competency_label_en",
            "level",
            "descriptor_en",
            "level_label_uk",
        ):
            if not cell(required):
                raise EmptyRequiredCellError(
                    f"sheet '{sheet_name}' row {row_idx}: empty required cell '{required}'"
                )
        level_str = cell("level")
        try:
            level = int(level_str)
        except ValueError as exc:
            raise EmptyRequiredCellError(
                f"sheet '{sheet_name}' row {row_idx}: level '{level_str}' is not an integer"
            ) from exc
        if not 1 <= level <= 5:
            raise EmptyRequiredCellError(
                f"sheet '{sheet_name}' row {row_idx}: level {level} is out of range 1..5"
            )

        block_slug = _slugify(block)
        comp = parsed.get(competency_id)
        if comp is None:
            comp = _ParsedCompetency(
                id=competency_id,
                block_slug=block_slug,
                block_label=block,
                label_uk=cell("competency_label_uk"),
                label_en=cell("competency_label_en"),
                topic=cell("topic"),
                retired=_norm(cell("competency_retired")).lower() == "true",
            )
            parsed[competency_id] = comp
        else:
            if comp.block_slug != block_slug:
                raise DuplicateCompetencyIdError(
                    f"sheet '{sheet_name}': competency_id '{competency_id}' appears in "
                    f"two blocks ('{comp.block_label}' and '{block}'); ids must be unique per sheet"
                )

        if level in comp.levels:
            raise DuplicateCompetencyIdError(
                f"sheet '{sheet_name}' row {row_idx}: competency_id '{competency_id}' "
                f"already declared level {level}"
            )
        evidence_raw = cell("evidence_examples")
        evidence = [e.strip() for e in evidence_raw.split(";") if e.strip()] if evidence_raw else []
        comp.levels[level] = _ParsedLevel(
            level=level,
            label_uk=cell("level_label_uk"),
            descriptor_en=cell("descriptor_en"),
            evidence_examples_en=evidence,
        )

    return parsed


def _competencies_to_yaml_doc(
    competencies: Iterable[_ParsedCompetency],
) -> dict[str, Any]:
    """Convert parsed competency records into the canonical YAML structure.

    Top-level (parent=null) nodes are one per unique block; competency nodes are
    children of those blocks. Each competency carries its `levels` array.
    """
    blocks: dict[str, str] = {}  # block_slug → block_label
    nodes: list[dict[str, Any]] = []
    for comp in competencies:
        if comp.block_slug not in blocks:
            blocks[comp.block_slug] = comp.block_label
            nodes.append(
                {
                    "id": f"block.{comp.block_slug}",
                    "label_en": comp.block_label,
                    "label_uk": comp.block_label,
                    "parent": None,
                    "retired": False,
                }
            )
        comp_node: dict[str, Any] = {
            "id": comp.id,
            "label_en": comp.label_en,
            "label_uk": comp.label_uk,
            "parent": f"block.{comp.block_slug}",
            "retired": comp.retired,
        }
        if comp.levels:
            comp_node["levels"] = [
                {
                    "level": lvl.level,
                    "label_uk": lvl.label_uk,
                    "descriptor_en": lvl.descriptor_en,
                    **(
                        {"evidence_examples_en": lvl.evidence_examples_en}
                        if lvl.evidence_examples_en
                        else {}
                    ),
                }
                for lvl in comp.levels.values()
            ]
        nodes.append(comp_node)
    return {"version": 1, "retired": False, "nodes": nodes}


# ---------------------------------------------------------------------------
# RubricImporter
# ---------------------------------------------------------------------------


class RubricImporter:
    """Convert + seed engine for the rubric matrix pipeline (T08)."""

    def __init__(self, schema_path: Path = _SCHEMA_PATH) -> None:
        self._schema = _load_schema(schema_path)

    # ---------- convert ----------

    def convert(self, xlsx_path: Path, out_dir: Path) -> list[Path]:
        """Read ``xlsx_path``, validate, and write canonical YAMLs to ``out_dir``.

        Returns the list of paths written, in canonical (sorted-by-filename) order.
        """
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=True, read_only=False)
        out_dir.mkdir(parents=True, exist_ok=True)
        written: list[Path] = []
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_"):
                continue
            stack_id = _norm(sheet_name)
            if not _STABLE_ID_RE.match(stack_id):
                raise EmptyRequiredCellError(
                    f"sheet '{sheet_name}': sheet name '{stack_id}' is not a valid stable id"
                )
            parsed = _read_sheet(wb[sheet_name], sheet_name)
            doc = _competencies_to_yaml_doc(parsed.values())
            _validate_yaml(doc, self._schema, origin=f"sheet '{sheet_name}'")
            out_path = out_dir / f"{stack_id}.yaml"
            out_path.write_text(_emit_canonical_yaml(doc), encoding="utf-8")
            written.append(out_path)
        return sorted(written)

    # ---------- seed ----------

    async def seed(
        self,
        yaml_dir: Path,
        *,
        dsn: str,
        dry_run: bool = False,
    ) -> SeedResult:
        """Reconcile the database with ``yaml_dir``. Idempotent on unchanged YAML.

        On payload-hash divergence: a new ``rubric_tree_version`` is created,
        a fresh tree is materialised under it, and one ``audit_log`` row is
        written. Prior-version rows are NEVER touched (§4).

        Rejects renames (FR-009) before any write — see :class:`RenameForbiddenError`.
        """
        docs = self._load_yaml_dir(yaml_dir)
        new_hash = _compute_payload_hash(yaml_dir)

        conn = await asyncpg.connect(_asyncpg_dsn(dsn))
        try:
            async with conn.transaction():
                await conn.execute("SELECT pg_advisory_xact_lock($1)", _SEED_ADVISORY_LOCK_ID)
                latest = await conn.fetchrow(
                    "SELECT id, payload_hash FROM rubric_tree_version "
                    "ORDER BY created_at DESC LIMIT 1"
                )
                if latest is not None and latest["payload_hash"] == new_hash:
                    return SeedResult(
                        noop=True,
                        new_version_id=None,
                        new_payload_hash=new_hash,
                        rows_inserted=0,
                    )
                # Rename detection (research §5).
                prior_active: set[str] = set()
                if latest is not None:
                    prior_rows = await conn.fetch(
                        "SELECT name FROM competency WHERE rubric_tree_version_id = $1",
                        latest["id"],
                    )
                    prior_active = {r["name"] for r in prior_rows}
                new_active: set[str] = set()
                new_retired: set[str] = set()
                for _, doc in docs:
                    for node in doc.get("nodes") or []:
                        node_id = node["id"]
                        if node_id.startswith("block."):
                            continue
                        if node.get("retired", False):
                            new_retired.add(node_id)
                        else:
                            new_active.add(node_id)
                disappeared = prior_active - new_active - new_retired
                if disappeared:
                    raise RenameForbiddenError(
                        f"stable id(s) {sorted(disappeared)} disappeared from the new payload "
                        f"without being retired. Renames are forbidden — retire the old id "
                        f"(set retired: true) and introduce the new id as a separate change."
                    )

                if dry_run:
                    return SeedResult(
                        noop=False,
                        new_version_id=None,
                        new_payload_hash=new_hash,
                        rows_inserted=0,
                    )

                # Insert the new rubric_tree_version row.
                version_row = await conn.fetchrow(
                    "INSERT INTO rubric_tree_version (label, is_active, payload_hash) "
                    "VALUES ($1, $2, $3) RETURNING id",
                    f"tree:{new_hash[:8]}",
                    True,
                    new_hash,
                )
                assert version_row is not None
                version_id = cast("UUID", version_row["id"])
                rows_inserted = 1

                # Materialise the tree per file (filename = stack name).
                for path, doc in docs:
                    rows_inserted += await self._materialise_stack(conn, version_id, path.stem, doc)

                # Audit row (FR-010).
                await conn.execute(
                    "INSERT INTO audit_log (actor_id, action, subject_hash) VALUES (NULL, $1, $2)",
                    "rubric.versioned",
                    new_hash,
                )
                rows_inserted += 1
                return SeedResult(
                    noop=False,
                    new_version_id=version_id,
                    new_payload_hash=new_hash,
                    rows_inserted=rows_inserted,
                )
        finally:
            await conn.close()

    # ---------- internals ----------

    def _load_yaml_dir(self, yaml_dir: Path) -> list[tuple[Path, dict[str, Any]]]:
        out: list[tuple[Path, dict[str, Any]]] = []
        for path in sorted(yaml_dir.glob("*.yaml")):
            doc = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            if not isinstance(doc, dict):
                raise SchemaViolationError(f"{path}: expected a mapping at top level")
            _validate_yaml(doc, self._schema, origin=str(path))
            out.append((path, doc))
        return out

    async def _materialise_stack(
        self,
        conn: asyncpg.Connection,
        version_id: UUID,
        stack_name: str,
        doc: dict[str, Any],
    ) -> int:
        """Insert one stack and its descendants. Returns count of rows written."""
        stack_row = await conn.fetchrow(
            "INSERT INTO stack (rubric_tree_version_id, name) VALUES ($1, $2) RETURNING id",
            version_id,
            stack_name,
        )
        assert stack_row is not None
        stack_id = cast("UUID", stack_row["id"])
        inserted = 1

        # Resolve nodes by id; assign each to its DB table by structural role.
        nodes_by_id: dict[str, dict[str, Any]] = {n["id"]: n for n in (doc.get("nodes") or [])}
        block_ids: dict[str, UUID] = {}  # stable_id → competency_block.id
        competency_ids: dict[str, UUID] = {}  # stable_id → competency.id

        # Pass 1: blocks (parent: null).
        for node in nodes_by_id.values():
            if node.get("parent") is None:
                block_row = await conn.fetchrow(
                    "INSERT INTO competency_block (rubric_tree_version_id, stack_id, name) "
                    "VALUES ($1, $2, $3) RETURNING id",
                    version_id,
                    stack_id,
                    node["id"],
                )
                assert block_row is not None
                block_ids[node["id"]] = cast("UUID", block_row["id"])
                inserted += 1

        # Pass 2: competencies (parent = a block id; have levels OR are leaf children of blocks).
        for node in nodes_by_id.values():
            parent_id = node.get("parent")
            if parent_id is not None and parent_id in block_ids:
                comp_row = await conn.fetchrow(
                    "INSERT INTO competency "
                    "(rubric_tree_version_id, competency_block_id, name) "
                    "VALUES ($1, $2, $3) RETURNING id",
                    version_id,
                    block_ids[parent_id],
                    node["id"],
                )
                assert comp_row is not None
                competency_ids[node["id"]] = cast("UUID", comp_row["id"])
                inserted += 1
                # Levels.
                for lvl in node.get("levels") or []:
                    await conn.execute(
                        "INSERT INTO level (rubric_tree_version_id, competency_id, "
                        "rank, descriptor) VALUES ($1, $2, $3, $4)",
                        version_id,
                        comp_row["id"],
                        lvl["level"],
                        lvl["descriptor_en"],
                    )
                    inserted += 1

        # Pass 3: topics (parent = a competency id).
        for node in nodes_by_id.values():
            parent_id = node.get("parent")
            if parent_id is not None and parent_id in competency_ids:
                await conn.execute(
                    "INSERT INTO topic (rubric_tree_version_id, competency_id, name) "
                    "VALUES ($1, $2, $3)",
                    version_id,
                    competency_ids[parent_id],
                    node["id"],
                )
                inserted += 1

        return inserted
